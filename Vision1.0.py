'''
Code Auther : Shoeb Ahmad (sahmad@provenconsult.com)
Project Name : Vision 1.0
Project Description : Computer vision based script that detect drowsiness and give alert
code technology : Python3.6
OS : Lunux (Ubunto)_
'''

# Library import

from scipy.spatial import distance
from playsound import playsound
from imutils import face_utils
from threading import Thread
import smtplib
import imutils
from openpyxl import *
from tkinter import *
import dlib
import cv2
import time

# DataBase func

def DataBase():
    # globally declare wb and sheet variable

    # opening the existing excel file
    wb = load_workbook('DataBase.xlsx')

    # create the sheet object
    sheet = wb.active


    def excel():
        # resize the width of columns in
        # excel spreadsheet
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 50

        # write given data to an excel spreadsheet
        # at particular location
        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "Company"
        sheet.cell(row=1, column=3).value = "Admin_ID"
        sheet.cell(row=1, column=4).value = "Employee ID"
        sheet.cell(row=1, column=5).value = "Contact Number"
        sheet.cell(row=1, column=6).value = "Email id"
        sheet.cell(row=1, column=7).value = "Location"


    # Function to set focus (cursor)
    def focus1(event):
        # set focus on the Company_field box
        Company_field.focus_set()


    # Function to set focus
    def focus2(event):
        # set focus on the Admin_ID_field box
        Admin_ID_field.focus_set()


    # Function to set focus
    def focus3(event):
        # set focus on the Employee_ID_field box
        Employee_ID_field.focus_set()


    # Function to set focus
    def focus4(event):
        # set focus on the contact_no_field box
        contact_no_field.focus_set()


    # Function to set focus
    def focus5(event):
        # set focus on the email_id_field box
        email_id_field.focus_set()


    # Function to set focus
    def focus6(event):
        # set focus on the Location_field box
        Location_field.focus_set()


    # Function for clearing the
    # contents of text entry boxes
    def clear():
        # clear the content of text entry box
        Name_field.delete(0, END)
        Company_field.delete(0, END)
        Admin_ID_field.delete(0, END)
        Employee_ID_field.delete(0, END)
        contact_no_field.delete(0, END)
        email_id_field.delete(0, END)
        Location_field.delete(0, END)


    # Function to take data from GUI
    # window and write to an excel file
    def insert():
        # if user not fill any entry
        # then print "empty input"
        if (Name_field.get() == "" and
                Company_field.get() == "" and
                Admin_ID_field.get() == "" and
                Employee_ID_field.get() == "" and
                contact_no_field.get() == "" and
                email_id_field.get() == "" and
                Location_field.get() == ""):

            print("empty input")

        else:

            # assigning the max row and max column
            # value upto which data is written
            # in an excel sheet to the variable
            current_row = sheet.max_row
            current_column = sheet.max_column

            # get method returns current text
            # as string which we write into
            # excel spreadsheet at particular location
            sheet.cell(row=current_row + 1, column=1).value = Name_field.get()
            sheet.cell(row=current_row + 1, column=2).value = Company_field.get()
            sheet.cell(row=current_row + 1, column=3).value = Admin_ID_field.get()
            sheet.cell(row=current_row + 1, column=4).value = Employee_ID_field.get()
            sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
            sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
            sheet.cell(row=current_row + 1, column=7).value = Location_field.get()

            # save the file
            wb.save('DataBase.xlsx')

            # set focus on the Name_field box
            Name_field.focus_set()

            # call the clear() function
            clear()

        # Driver code


    if __name__ == "__main__":
        # create a GUI window
        root = Tk()

        # set the background colour of GUI window
        root.configure(background='light green')

        # set the title of GUI window
        root.title("registration form")

        # set the configuration of GUI window
        root.geometry("500x300")

        excel()

        # create a Form label
        heading = Label(root, text="Form", bg="light green")

        # create a Name label
        Name = Label(root, text="Name", bg="light green")

        # create a Company label
        Company = Label(root, text="Company", bg="light green")

        # create a Admin_IDester label
        Admin_ID = Label(root, text="Admin_ID", bg="light green")

        # create a Form No. lable
        Employee_ID = Label(root, text="Employee_ID", bg="light green")

        # create a Contact No. label
        contact_no = Label(root, text="Contact No.", bg="light green")

        # create a Email id label
        email_id = Label(root, text="Email id", bg="light green")

        # create a Location label
        Location = Label(root, text="Location", bg="light green")

        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        heading.grid(row=0, column=1)
        Name.grid(row=1, column=0)
        Company.grid(row=2, column=0)
        Admin_ID.grid(row=3, column=0)
        Employee_ID.grid(row=4, column=0)
        contact_no.grid(row=5, column=0)
        email_id.grid(row=6, column=0)
        Location.grid(row=7, column=0)

        # create a text entry box
        # for typing the information
        Name_field = Entry(root)
        Company_field = Entry(root)
        Admin_ID_field = Entry(root)
        Employee_ID_field = Entry(root)
        contact_no_field = Entry(root)
        email_id_field = Entry(root)
        Location_field = Entry(root)

        # bind method of widget is used for
        # the binding the function with the events

        # whenever the enter key is pressed
        # then call the focus1 function
        Name_field.bind("<Return>", focus1)

        # whenever the enter key is pressed
        # then call the focus2 function
        Company_field.bind("<Return>", focus2)

        # whenever the enter key is pressed
        # then call the focus3 function
        Admin_ID_field.bind("<Return>", focus3)

        # whenever the enter key is pressed
        # then call the focus4 function
        Employee_ID_field.bind("<Return>", focus4)

        # whenever the enter key is pressed
        # then call the focus5 function
        contact_no_field.bind("<Return>", focus5)

        # whenever the enter key is pressed
        # then call the focus6 function
        email_id_field.bind("<Return>", focus6)

        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        Name_field.grid(row=1, column=1, ipadx="100")
        Company_field.grid(row=2, column=1, ipadx="100")
        Admin_ID_field.grid(row=3, column=1, ipadx="100")
        Employee_ID_field.grid(row=4, column=1, ipadx="100")
        contact_no_field.grid(row=5, column=1, ipadx="100")
        email_id_field.grid(row=6, column=1, ipadx="100")
        Location_field.grid(row=7, column=1, ipadx="100")

        # call excel function
        excel()

        # create a Submit Button and place into the root window
        submit = Button(root, text="Submit", fg="Black",
                        bg="Red", command=insert)
        submit.grid(row=8, column=1)

        # start the GUI
        root.mainloop()
        return

DataBase()

# Mail function

def Mail():
    try:
        gmailaddress = ("er.shoaib10@gmail.com")
        gmailpassword = ("1016331062")
        mailto = ("goodthink9211@gmail.com")
        subj = ("Vision 1.0 Generated mail, Please do not reply!")
        msg = ("Hi Admin , Pleased to notify you that"+" "+Name+" "+"is sleeping in office time")
        mailServer = smtplib.SMTP('smtp.gmail.com' , 587)
        mailServer.starttls()
        mailServer.login(gmailaddress , gmailpassword)
        mailServer.sendmail(gmailaddress, mailto , msg)
        print(" \n  Mail Sent!")
        mailServer.quit()
    except ConnectionError as e:
        print(e)

# Sound Alert

def Alert():
    playsound('Siren.wav')

    return

# Eye spacing calculation

def eye_aspect_ratio(eye):
    A = distance.euclidean(eye[1], eye[5])
    B = distance.euclidean(eye[2], eye[4])
    C = distance.euclidean(eye[0], eye[3])
    ear = (A + B) / (2.0 * C)
    return ear

# Configuration Elements based on requirement

thresh = 0.27
frame_check = 30
detect = dlib.get_frontal_face_detector()
ALARM_ON = False
predict = dlib.shape_predictor('shape_predictor_68_face_landmarks.dat')  # Dat file is the crux of the code

(lStart, lEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["left_eye"]
(rStart, rEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["right_eye"]
cap = cv2.VideoCapture(0)
flag = 0

# Main looping for camera detection and fall back

while True:

    ret, frame = cap.read()
    frame = imutils.resize(frame, width=450)
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    subjects = detect(gray, 0)

    for subject in subjects:

        shape = predict(gray, subject)
        shape = face_utils.shape_to_np(shape)  # converting to NumPy Array
        leftEye = shape[lStart:lEnd]
        rightEye = shape[rStart:rEnd]
        leftEAR = eye_aspect_ratio(leftEye)
        rightEAR = eye_aspect_ratio(rightEye)
        ear = (leftEAR + rightEAR) / 2.0
        leftEyeHull = cv2.convexHull(leftEye)
        rightEyeHull = cv2.convexHull(rightEye)
        cv2.drawContours(frame, [leftEyeHull], -1, (0, 255, 0), 1)
        cv2.drawContours(frame, [rightEyeHull], -1, (0, 255, 0), 1)

        if ear < thresh:

            flag += 1
            print(flag)

            # Alert Sound

            if flag >= frame_check:
                if not ALARM_ON:
                    ALARM_ON = True
                    t = Thread(target=Alert(),
                               args=(["alarm"],))
                    t.deamon = True




                cv2.putText(frame, "****************ALERT!****************", (10, 50),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                cv2.putText(frame, "****************ALERT!****************", (10, 325),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)

            # print ("Drowsy")
            #Alert()

        else:

            flag = 0
            ALARM_ON = False

    cv2.imshow("Frame", frame)
    key = cv2.waitKey(1) & 0xFF

    if key == ord("q"):
        break

cv2.destroyAllWindows()
cap.stop()